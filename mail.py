"""
EPİAŞ Şeffaflık Platformu - Kesinleşmemiş PFT Mail Gönderici
--------------------------------------------------------------
Kurulum:  uv pip install eptr2

Cron job (her gün 13:00'de):
  0 13 * * * python epias_pft_mail.py >> epias_pft.log 2>&1

TEST_MODU = True  → dünün verisini çeker (deneme)
TEST_MODU = False → yarının verisini çeker (normal çalışma)
"""

import smtplib
import logging
import io
import base64
from datetime import datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList

from eptr2 import EPTR2

# ─────────────────────────────────────────────
#  AYARLAR — Buraya kendi bilgilerini gir
# ─────────────────────────────────────────────

EPIAS_USERNAME = "beyzayn1168@gmail.com"
EPIAS_PASSWORD = "5630Byz."

import os
EPIAS_USERNAME = os.environ["EPIAS_USERNAME"]
EPIAS_PASSWORD = os.environ["EPIAS_PASSWORD"]
OUTLOOK_MAIL   = os.environ["OUTLOOK_MAIL"]
OUTLOOK_PASS   = os.environ["OUTLOOK_PASS"]

SMTP_SERVER  = "smtp.office365.com"
SMTP_PORT    = 587

# True = dünün verisi (test), False = yarının verisi (normal)
TEST_MODU = True

MUSTERI_LISTESI = [
    {"ad": "Müşteri Bir", "email": "beyzanur.ozbek@alpineenerji.com.tr"}
]

#  LOGGING
# ─────────────────────────────────────────────
#  LOGGING
# ─────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger(__name__)


# ─────────────────────────────────────────────
#  YARDIMCI: saat aralığı formatla
# ─────────────────────────────────────────────

def saat_aralik(saat_no: str) -> str:
    """'1' → '01:00-02:00', '0' → '00:00-01:00'"""
    try:
        h = int(saat_no)
        return f"{h:02d}:00-{(h + 1) % 24:02d}:00"
    except Exception:
        return saat_no


# ─────────────────────────────────────────────
#  PFT VERİSİ ÇEK
# ─────────────────────────────────────────────
# --- AYARLAR ---
# Artık TEST_MODU değişkenine gerek yok, doğrudan yarını hedefliyoruz.

def pft_veri_cek():
    # Yarının tarihini hesapla
    hedef = datetime.now() + timedelta(days=1)
    tarih_str = hedef.strftime("%Y-%m-%d")

    log.info(f"Yarının kesinleşmemiş PFT verisi çekiliyor: {tarih_str}")

    eptr = EPTR2(username=EPIAS_USERNAME, password=EPIAS_PASSWORD)
    df = eptr.call("interim-mcp", start_date=tarih_str, end_date=tarih_str)

    veri = []
    for _, row in df.iterrows():
        saat_no = str(row.get("hour", ""))
        fiyat = row.get("marketTradePrice", None)
        veri.append({
            "saat_no": saat_no,
            "saat": saat_aralik(saat_no),
            "fiyat": fiyat if fiyat is not None else 0.0,
            "fiyat_str": str(fiyat) if fiyat is not None else "-",
        })

    if not veri:
        raise ValueError(f"{tarih_str} tarihli PFT verisi henüz yayınlanmamış.")

    log.info(f"{len(veri)} saatlik kayıt alındı.")
    return veri, tarih_str


def grafik_olustur(veri: list, tarih: str) -> str:
    import matplotlib.pyplot as plt
    import matplotlib.ticker as mticker
    import io, base64

    def format_aralik(s_no):
        try:
            h = int(s_no.split(":")[0])
        except:
            h = int(s_no)
        return f"{h:02d}:00-{(h + 1) % 24:02d}:00"

    saat_araliklari = [format_aralik(r["saat_no"]) for r in veri]
    fiyatlar = [float(r["fiyat"]) for r in veri]
    tarih_fmt = datetime.strptime(tarih, "%Y-%m-%d").strftime("%d.%m.%Y")

    NAVY = "#201F5A"
    BRIGHT_NAVY = "#2b2982"

    # Genişliği 14 inç yaparak tablo genişliğiyle (1300px) senkronize ettik
    fig, ax = plt.subplots(figsize=(14, 5.5))
    fig.patch.set_facecolor("white")

    ax.plot(range(len(saat_araliklari)), fiyatlar, color=NAVY, linewidth=2.5, marker='o', markersize=5)
    ax.fill_between(range(len(saat_araliklari)), fiyatlar, color=NAVY, alpha=0.08)

    for spine in ax.spines.values():
        spine.set_visible(True)
        spine.set_edgecolor('#BBBBBB')
        spine.set_linewidth(1.2)

    ax.set_xticks(range(len(saat_araliklari)))
    ax.set_xticklabels(saat_araliklari, rotation=45, ha="right", fontsize=9)
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: "{:,.0f}".format(x)))
    ax.grid(axis="y", linestyle="--", alpha=0.3)

    fig.text(0.98, 0.92, "ALPİNE", fontsize=13, fontweight="black", color=BRIGHT_NAVY, ha="right")
    fig.text(0.98, 0.85, "ENERJİ", fontsize=13, fontweight="black", color=BRIGHT_NAVY, ha="right")

    plt.title(f"EPİAŞ Kesinleşmemiş PFT - {tarih_fmt}", fontsize=13, fontweight="bold", color="#222", pad=35)

    # Sağdaki boşluğu sıfırlayarak tablonun sonuyla hizaladık
    plt.tight_layout(rect=[0, 0, 1, 0.95])

    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=110, bbox_inches="tight")
    plt.close(fig)
    return base64.b64encode(buf.getvalue()).decode("utf-8")
# ─────────────────────────────────────────────
#  XLSX OLUŞTUR
# ─────────────────────────────────────────────
def xlsx_olustur(veri: list, tarih: str) -> bytes:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt, matplotlib.ticker as mticker, io as _io
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    tarih_fmt = datetime.strptime(tarih, "%Y-%m-%d").strftime("%d.%m.%Y")

    # Renk Kodları
    NAVY_HEX = "201F5A"
    NAVY_GRAPHIC = "#201F5A"

    # --- STİLLER ---
    header_fill = PatternFill("solid", start_color=NAVY_HEX, end_color=NAVY_HEX)
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)

    # Metni Kaydır (Wrap Text) özelliğini buraya ekledik
    wrap_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center")

    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"), bottom=Side(style="thin", color="CCCCCC"),
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PFT Verileri"

    # Sütun Genişlikleri
    for col, w in zip("ABC", [15, 20, 20]):
        ws.column_dimensions[col].width = w

    # --- EXCEL ÜST BAŞLIK ---
    ws.row_dimensions[1].height = 45  # Yüksekliği biraz artırdık ki kayan metin sığsın

    ws.merge_cells("A1:B1")
    ws["A1"] = "ALPİNE ENERJİ"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color=NAVY_HEX)
    ws["A1"].alignment = center_align

    # Başlık Hücresi: Metni Kaydır aktif edildi
    ws["C1"] = f"EPİAŞ Kesinleşmemiş PFT - {tarih_fmt}"
    ws["C1"].font = Font(name="Arial", size=11, bold=True, color=NAVY_HEX)
    ws["C1"].alignment = wrap_alignment

    # --- TABLO BAŞLIKLARI (Satır 4) ---
    ws.row_dimensions[4].height = 22
    for col_idx, h in enumerate(["Tarih", "Saat Aralığı", "PFT (TL/MWh)"], start=1):
        c = ws.cell(row=4, column=col_idx, value=h)
        c.font = header_font;
        c.fill = header_fill;
        c.alignment = center_align;
        c.border = thin_border

    # --- VERİ SATIRLARI (Satır 5'ten başlar) ---
    for i, row in enumerate(veri, start=5):
        fiyat = float(row["fiyat"])
        bg = "EFF4FB" if i % 2 == 0 else "FFFFFF"
        rf = PatternFill("solid", start_color=bg, end_color=bg)
        bold_font = Font(name="Arial", size=10, bold=True, color="000000")
        normal_font = Font(name="Arial", size=10, color="000000")

        c = ws.cell(row=i, column=1, value=tarih_fmt);
        c.font = normal_font;
        c.alignment = center_align;
        c.fill = rf;
        c.border = thin_border
        c = ws.cell(row=i, column=2, value=row["saat"]);
        c.font = bold_font;
        c.alignment = center_align;
        c.fill = rf;
        c.border = thin_border

        c = ws.cell(row=i, column=3, value=fiyat)
        c.font = bold_font;
        c.number_format = '#,##0.00';
        c.alignment = center_align;
        c.fill = rf;
        c.border = thin_border

    # --- MATPLOTLIB GRAFİK ---
    saatler = [r["saat_no"] + ":00" for r in veri]
    fiyatlar = [float(r["fiyat"]) for r in veri]

    fig, ax = plt.subplots(figsize=(8, 4))
    fig.patch.set_facecolor("white")

    ax.bar(range(len(saatler)), fiyatlar, color=NAVY_GRAPHIC, width=0.7)

    ax.set_xticks(range(len(saatler)))
    ax.set_xticklabels(saatler, rotation=45, ha="right", fontsize=8)
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: "{:,.0f}".format(x)))
    ax.grid(axis="y", linestyle="--", alpha=0.3)

    # SAĞ ÜST KÖŞE LOGO (Grafik alanı dışında)
    fig.text(0.95, 0.92, "ALPİNE", fontsize=11, fontweight="black", color=NAVY_GRAPHIC, ha="right", va="top")
    fig.text(0.95, 0.86, "ENERJİ", fontsize=11, fontweight="black", color=NAVY_GRAPHIC, ha="right", va="top")

    # ORTALANMIŞ BAŞLIK
    plt.title(f"EPİAŞ Kesinleşmemiş PFT - {tarih_fmt}",
              fontsize=11, loc='center', color="#222222", fontweight="bold", pad=25)

    plt.tight_layout(rect=[0, 0, 0.9, 0.95])

    img_buf = _io.BytesIO()
    fig.savefig(img_buf, format="png", dpi=100, bbox_inches="tight")
    plt.close(fig)
    img_buf.seek(0)

    # --- GRAFİĞİ E5 HÜCRESİNE YERLEŞTİR ---
    from openpyxl.drawing.image import Image as XLImage
    xl_img = XLImage(img_buf)
    xl_img.anchor = "E5"
    ws.add_image(xl_img)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
# ─────────────────────────────────────────────
#  HTML MAIL (grafik gömülü + yatay tablo)
# ─────────────────────────────────────────────
def html_mail_olustur(musteri_ad: str, veri: list, tarih: str, grafik_b64: str) -> str:
    tarih_fmt = datetime.strptime(tarih, "%Y-%m-%d").strftime("%d.%m.%Y")
    NAVY = "#201F5A"

    def format_aralik(s_no):
        try: h = int(s_no.split(":")[0])
        except: h = int(s_no)
        return f"{h:02d}:00-{(h+1)%24:02d}:00"

    th_cells = "".join(
        f"<th style='padding:10px 4px; background:{NAVY}; color:#fff; font-size:10px; border:1px solid #ddd; min-width:95px;'>{format_aralik(r['saat_no'])}</th>"
        for r in veri
    )

    # Fiyat puntosu 13px yapıldı ve hücre içi ferahlatıldı
    td_cells = "".join(
        f"<td style='padding:12px 4px; font-size:13px; text-align:center; font-weight:bold; border:1px solid #ddd; color:{NAVY};'>{float(r['fiyat']):.2f}</td>"
        for r in veri
    )

    return f"""
    <html>
    <body style="font-family: Arial, sans-serif; color: #222; margin:0; padding:0;">
      <table width="100%" style="border-collapse:collapse; background:{NAVY}; color:#fff;">
        <tr>
          <td style="padding:25px;">
            <div style="font-size:24px; font-weight:bold;">ALPİNE ENERJİ</div>
            <div style="font-size:14px; color:#4EB2D2; margin-top:5px;">{tarih_fmt} Tarihine Ait Kesinleşmemiş PFT</div>
          </td>
        </tr>
      </table>

      <div style="padding:30px; border:1px solid #ddd; border-top:none;">
        <p style="font-size:15px;">Sayın <b>{musteri_ad}</b>,</p>
        <p style="font-size:15px;">{tarih_fmt} tarihine ait <b>kesinleşmemiş Piyasa Fiyatı Tahmini (PFT)</b> verileri aşağıda yer almaktadır.</p>

        <!-- GENİŞLİK AYARLANMIŞ GRAFİK -->
        <div style="margin:25px 0; width:1300px;">
          <img src="data:image/png;base64,{grafik_b64}" style="width:1300px; display:block; border-radius:4px;" />
        </div>

        <!-- GENİŞLİK AYARLANMIŞ TABLO -->
        <div style="overflow-x:auto;">
          <table style="border-collapse:collapse; width:1300px; table-layout: fixed; border:1px solid #ddd;">
            <thead>
              <tr>
                <th style="padding:12px; background:{NAVY}; color:#fff; font-size:11px; border:1px solid #ddd; text-align:left; width:140px;">Saat Aralığı</th>
                {th_cells}
              </tr>
            </thead>
            <tbody>
              <tr>
                <td style="padding:12px; background:{NAVY}; color:#fff; font-size:11px; border:1px solid #ddd; font-weight:bold;">PFT (TL/MWh)</td>
                {td_cells}
              </tr>
            </tbody>
          </table>
        </div>

        <p style="font-size:12px; color:#666; margin-top:35px; border-top:1px solid #eee; padding-top:15px;">
          ⚠️ Bu veriler <u>kesinleşmemiş</u> olup değişiklik gösterebilir. Detaylı analizler ekteki Excel dosyasındadır.
        </p>
      </div>
    </body>
    </html>
    """
# ─────────────────────────────────────────────
#  MAIL GÖNDER
# ─────────────────────────────────────────────

def mail_gonder(musteri: dict, veri: list, tarih: str, xlsx_bytes: bytes, grafik_b64: str):
    ad = musteri["ad"]
    email = musteri["email"]

    tarih_fmt = datetime.strptime(tarih, "%Y-%m-%d").strftime("%d.%m.%Y")
    konu = (
        f"[TEST] EPİAŞ Kesinleşmemiş PFT — {tarih_fmt}"
        if TEST_MODU else
        f"EPİAŞ Kesinleşmemiş PFT — {tarih_fmt}"
    )

    msg = MIMEMultipart("mixed")
    msg["Subject"] = konu
    msg["From"] = OUTLOOK_MAIL
    msg["To"] = email

    html_icerik = html_mail_olustur(ad, veri, tarih, grafik_b64)
    msg.attach(MIMEText(html_icerik, "html", "utf-8"))

    # XLSX eki
    ek = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    ek.set_payload(xlsx_bytes)
    encoders.encode_base64(ek)
    ek.add_header("Content-Disposition", f'attachment; filename="PFT_{tarih}.xlsx"')
    msg.attach(ek)

    with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
        server.ehlo()
        server.starttls()
        server.login(OUTLOOK_MAIL, OUTLOOK_PASS)
        server.sendmail(OUTLOOK_MAIL, email, msg.as_string())

    log.info(f"✓ Mail gönderildi → {email}")


# ─────────────────────────────────────────────
#  ANA AKIŞ
# ─────────────────────────────────────────────

def main():
    log.info("=" * 55)
    log.info(f"Başlatıldı. Mod: {'TEST (dün)' if TEST_MODU else 'NORMAL (yarın)'}")

    try:
        veri, tarih = pft_veri_cek()
        xlsx_bytes = xlsx_olustur(veri, tarih)
        grafik_b64 = grafik_olustur(veri, tarih)

        basarili = basarisiz = 0
        for musteri in MUSTERI_LISTESI:
            try:
                mail_gonder(musteri, veri, tarih, xlsx_bytes, grafik_b64)
                basarili += 1
            except Exception as e:
                log.error(f"✗ Mail gönderilemedi → {musteri['email']}: {e}")
                basarisiz += 1

        log.info(f"Tamamlandı. ✓ {basarili} gönderildi, ✗ {basarisiz} hata.")

    except Exception as e:
        log.error(f"KRİTİK HATA: {e}")
        raise


if __name__ == "__main__":
    main()
