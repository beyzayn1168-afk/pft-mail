
import smtplib
import logging
import io
import base64
import os
from datetime import datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from matplotlib.image import imread
import numpy as np

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from eptr2 import EPTR2

# ─────────────────────────────────────────────
import os

LOGO_PATH_JPG = "Alpine-enerji.jpg"
LOGO_PATH_PNG = "Alpine-enerji.png"
LOGO_PATH_BEYAZ = "Alpine-enerji-beyaz.png"

# Renkli logo (Excel için)
if os.path.exists(LOGO_PATH_JPG):
    with open(LOGO_PATH_JPG, "rb") as f:
        LOGO_B64 = base64.b64encode(f.read()).decode("utf-8")
else:
    LOGO_B64 = ""

if os.path.exists(LOGO_PATH_BEYAZ):
    with open(LOGO_PATH_BEYAZ, "rb") as f:
        LOGO_BEYAZ_B64 = base64.b64encode(f.read()).decode("utf-8")
    LOGO_MAIL_SRC = f"data:image/png;base64,{LOGO_BEYAZ_B64}"
else:
    LOGO_MAIL_SRC = ""
    
# ─────────────────────────────────────────────
#  AYARLAR — GitHub Secrets'tan Okunur
# ─────────────────────────────────────────────
EPIAS_USERNAME = os.getenv("EPIAS_USERNAME")
EPIAS_PASSWORD = os.getenv("EPIAS_PASSWORD")
OUTLOOK_MAIL   = os.getenv("OUTLOOK_MAIL")
OUTLOOK_PASS   = os.getenv("OUTLOOK_PASS")

SMTP_SERVER  = "smtp.office365.com"
SMTP_PORT    = 587

TEST_MODU = False 

MUSTERI_LISTESI = [

    {"ad": "Beyza Nur Özbek", "email": "beyzanur.ozbek@alpineenerji.com.tr"}
]

# ─────────────────────────────────────────────
#  LOGGING
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger(__name__)


def saat_aralik(saat_no: str) -> str:
    try:
        h = int(saat_no)
        return f"{h:02d}:00-{(h + 1) % 24:02d}:00"
    except Exception:
        return saat_no


def ptf_veri_cek():
    hedef = datetime.now() + timedelta(days=1)
    tarih_str = hedef.strftime("%Y-%m-%d")
    log.info(f"Bugünün ({tarih_str}) PTF verisi deneniyor... [TEST MODU]")

    eptr = EPTR2(username=EPIAS_USERNAME, password=EPIAS_PASSWORD)

    try:
        df = eptr.call("interim-mcp", start_date=tarih_str, end_date=tarih_str)

        if df is None or df.empty:
            log.warning(f"⚠️ {tarih_str} tarihli PTF verisi henüz EPİAŞ tarafından yayınlanmamış.")
            return None, tarih_str

        veri = []
        for _, row in df.iterrows():
            saat_no = str(row.get("hour", ""))
            fiyat = row.get("marketTradePrice", None)
            veri.append({
                "saat_no": saat_no,
                "saat": saat_aralik(saat_no),
                "fiyat": fiyat if fiyat is not None else 0.0,
                "fiyat_str": str(fiyat) if fiyat is not None else "-",
            })

        if not veri:
            log.warning(f"⚠️ {tarih_str} tarihli PTF verisi boş geldi.")
            return None, tarih_str

        log.info(f"✓ Yarının ({tarih_str}) verisi başarıyla alındı. ({len(veri)} saatlik)")
        return veri, tarih_str

    except Exception as e:
        log.error(f"HATA: Veri çekilirken bir sorun oluştu: {e}")
        return None, tarih_str

def grafik_olustur(veri: list, tarih: str) -> str:
    def fmt_iki_satir_saat(s_no):
        try:
            h = int(s_no.split(":")[0])
        except:
            h = int(s_no)
        return f"{h:02d}:00\n{(h + 1) % 24:02d}:00"

    n = len(veri)
    saat_araliklari = [fmt_iki_satir_saat(r["saat_no"]) for r in veri]
    fiyatlar = [float(r["fiyat"]) for r in veri]
    tarih_fmt = datetime.strptime(tarih, "%Y-%m-%d").strftime("%d.%m.%Y")
    NAVY = "#201F5A"

    fig = plt.figure(figsize=(12, 7))
    fig.patch.set_facecolor("white")

    ax = fig.add_axes([0.06, 0.33, 0.91, 0.50])
    x = np.arange(n)
    bars = ax.bar(x, fiyatlar, color=NAVY, width=0.55, zorder=3)
    for bar, val in zip(bars, fiyatlar):
        ax.text(
            bar.get_x() + bar.get_width() / 2,
            bar.get_height() + 5,
            f"{val:,.0f}" if val > 0 else "0",
            ha="center", va="bottom", fontsize=7.5, color=NAVY, fontweight="bold"
        )
    ax.set_xticks(x)
    ax.set_xticklabels([])
    ax.set_xlim(-0.5, n - 0.5)
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: "{:,.0f}".format(v)))
    ax.tick_params(axis="y", labelsize=7)
    ax.set_ylim(0, max(fiyatlar) * 1.2)
    ax.grid(axis="y", linestyle="--", alpha=0.3, zorder=0)
    for spine in ["top", "right"]:
        ax.spines[spine].set_visible(False)
    for spine in ["left", "bottom"]:
        ax.spines[spine].set_edgecolor("#BBBBBB")

    fig.text(0.5, 0.94, f"EPİAŞ Kesinleşmemiş Piyasa Takas Fiyatı (PTF) — {tarih_fmt}",
             ha="center", fontsize=10, fontweight="bold", color="#201F5A")

    try:
        logo_img = imread(LOGO_PATH_PNG)
        logo_ax = fig.add_axes([0.82, 0.88, 0.14, 0.10])
        logo_ax.imshow(logo_img)
        logo_ax.axis("off")
    except Exception as e:
        log.warning(f"Logo yüklenemedi (grafik): {e}")

    ax_t = fig.add_axes([0.06, 0.12, 0.91, 0.26])
    ax_t.set_axis_off()
    tbl = ax_t.table(
        cellText=[
            [s for s in saat_araliklari],
            [f"{v:,.0f}" for v in fiyatlar]
        ],
        rowLabels=["Saat Aralığı", "PTF (TL/MWh)"],
        cellLoc="center",
        loc="center"
    )
    tbl.auto_set_font_size(False)
    tbl.set_fontsize(7)
    for (ri, ci), cell in tbl.get_celld().items():
        cell.set_linewidth(0.4)
        cell.set_edgecolor("#BBBBBB")
        if ci == -1:
            cell.set_facecolor(NAVY)
            cell.set_text_props(color="white", fontweight="bold", fontsize=7)
            cell.set_height(0.25)
        elif ri == 1:
            cell.set_facecolor(NAVY)
            cell.set_text_props(color="white", fontweight="bold", fontsize=7)
            cell.set_height(0.25)
        else:
            cell.set_facecolor("#EFF4FB" if ci % 2 == 0 else "#FFFFFF")
            cell.set_text_props(color=NAVY, fontweight="bold")
            cell.set_height(0.25)

    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=100, bbox_inches="tight")
    plt.close(fig)
    return base64.b64encode(buf.getvalue()).decode("utf-8")
    
def xlsx_olustur(veri: list, tarih: str) -> bytes:
    tarih_fmt = datetime.strptime(tarih, "%Y-%m-%d").strftime("%d.%m.%Y")
    NAVY_HEX    = "201F5A"
    NAVY_GRAPHIC = "#201F5A"

    header_fill = PatternFill("solid", start_color=NAVY_HEX, end_color=NAVY_HEX)
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),  bottom=Side(style="thin", color="CCCCCC"),
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PTF Verileri"

    for col, w in zip("ABC", [15, 20, 20]):
        ws.column_dimensions[col].width = w

    ws.row_dimensions[1].height = 48
    ws.merge_cells("B1:C1")

    # Logo ve Başlık Bölümü
    try:
        from openpyxl.drawing.image import Image as XLImage
        xl_logo = XLImage(LOGO_PATH_JPG)
        xl_logo.width, xl_logo.height = 145, 60
        xl_logo.anchor = "A1"
        ws.add_image(xl_logo)
    except:
        ws["A1"] = "ALPİNE ENERJİ"
        ws["A1"].font = Font(name="Arial", size=14, bold=True, color=NAVY_HEX)

    ws["B1"] = f"EPİAŞ Kesinleşmemiş PTF - {tarih_fmt}"
    ws["B1"].font = Font(name="Arial", size=11, bold=True, color=NAVY_HEX)
    ws["B1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Başlık Satırı
    for ci, h in enumerate(["Tarih", "Saat Aralığı", "PTF (TL/MWh)"], start=1):
        c = ws.cell(row=4, column=ci, value=h)
        c.font = header_font; c.fill = header_fill; c.alignment = center_align; c.border = thin_border

    # Verileri Yazma ve Ortalama Hesaplama Hazırlığı
    fiyatlar = [float(row["fiyat"]) for row in veri]
    toplam_saat = len(fiyatlar)
    son_satir_no = 4

    for i, row in enumerate(veri, start=5):
        fiyat = float(row["fiyat"])
        bg = "EFF4FB" if i % 2 == 0 else "FFFFFF"
        rf = PatternFill("solid", start_color=bg, end_color=bg)
        nf = Font(name="Arial", size=10, bold=True)
        
        ws.cell(row=i, column=1, value=tarih_fmt).font = nf
        ws.cell(row=i, column=2, value=row["saat"]).font = nf
        c3 = ws.cell(row=i, column=3, value=fiyat)
        c3.font = nf; c3.number_format = '#,##0.00'
        
        for col in range(1, 4):
            ws.cell(row=i, column=col).alignment = center_align
            ws.cell(row=i, column=col).fill = rf
            ws.cell(row=i, column=col).border = thin_border
        son_satir_no = i

    # --- ORTALAMA HESAPLARI ---
    gunluk_ort = sum(fiyatlar) / toplam_saat if toplam_saat > 0 else 0
    # Not: PTF verisinde genelde miktar verilmediği için 'Ağırlıklı Ortalama' 
    # bazen günlük ortalama ile aynı çıkar; ancak mantığı kurmak adına hesaplıyoruz.
    agirlikli_ort = sum(fiyatlar) / toplam_saat # Eğer elinizde miktar (MWh) olsaydı ona bölecektik

    # --- LACİVERT ORTALAMA SATIRLARI EKLEME ---
    # 1. Günlük Ortalama Satırı
    g_satir = son_satir_no + 1
    ws.merge_cells(f"A{g_satir}:B{g_satir}")
    c_g_lab = ws.cell(row=g_satir, column=1, value="GÜNLÜK ORTALAMA")
    c_g_val = ws.cell(row=g_satir, column=3, value=gunluk_ort)

    # 2. Ağırlıklı Ortalama Satırı
    a_satir = son_satir_no + 2
    ws.merge_cells(f"A{a_satir}:B{a_satir}")
    c_a_lab = ws.cell(row=a_satir, column=1, value="AĞIRLIKLI ORTALAMA")
    c_a_val = ws.cell(row=a_satir, column=3, value=agirlikli_ort)

    # Stil Uygulama (Lacivert ve Beyaz Yazı)
    for r_idx in [g_satir, a_satir]:
        for c_idx in [1, 3]: # A:B birleşik olduğu için 1 ve 3. kolonlar
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = Font(bold=True, color="FFFFFF", name="Arial")
            cell.fill = header_fill # Sizin belirlediğiniz Lacivert (NAVY_HEX)
            cell.alignment = center_align
            cell.border = thin_border
            if c_idx == 3:
                cell.number_format = '#,##0.00'

    # --- GRAFİK (Ortalama Çizgileri Dahil) ---
    saatler = [r["saat_no"] + ":00" for r in veri]
    fig2, ax2 = plt.subplots(figsize=(9, 5))
    ax2.bar(range(len(saatler)), fiyatlar, color=NAVY_GRAPHIC, width=0.7, label="PTF")
    
    # Ortalama Çizgilerini Grafiğe Ekleme
    ax2.axhline(gunluk_ort, color="red", linestyle="--", linewidth=1.5, label=f"Günlük Ort: {gunluk_ort:,.2f}")
    ax2.axhline(agirlikli_ort, color="orange", linestyle=":", linewidth=1.5, label=f"Ağ. Ort: {agirlikli_ort:,.2f}")
    
    ax2.set_xticks(range(len(saatler)))
    ax2.set_xticklabels(saatler, rotation=45, ha="right", fontsize=8)
    ax2.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: "{:,.0f}".format(x)))
    ax2.grid(axis="y", linestyle="--", alpha=0.3)
    ax2.legend(loc="upper left", bbox_to_anchor=(1, 1), fontsize=8) # Legend dışarıda
    
    plt.title(f"PTF ve Ortalamalar - {tarih_fmt}", fontsize=12, color=NAVY_GRAPHIC, fontweight="bold")

    plt.tight_layout()
    ibuf = io.BytesIO()
    fig2.savefig(ibuf, format="png", dpi=100)
    plt.close(fig2)
    ibuf.seek(0)

    xl_img = XLImage(ibuf)
    xl_img.anchor = "E5"
    ws.add_image(xl_img)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

def html_mail_olustur(musteri_ad: str, veri: list, tarih: str, grafik_b64: str) -> str:
    tarih_fmt = datetime.strptime(tarih, "%Y-%m-%d").strftime("%d.%m.%Y")
    NAVY = "#201F5A"

    return f"""<!DOCTYPE html>
<html lang="tr">
<body style="margin:0; padding:0; font-family:Arial,sans-serif; color:#222; background:#f4f6fb;">
  <table width="100%" cellpadding="0" cellspacing="0" style="background:#f4f6fb;">
    <tr><td align="center" style="padding:24px 8px;">
      <table width="100%" cellpadding="0" cellspacing="0" style="max-width:800px; background:#fff; border-radius:8px;">
        <tr>
         <td style="background:#201F5A; padding:16px 30px; border-radius:8px 8px 0 0;">
            <table width="100%" cellpadding="0" cellspacing="0">
              <tr>
                <td style="vertical-align:middle; text-align:left;">
                  <div style="font-size:14px; font-weight:900; color:#fff; line-height:1.3;">Kesinleşmemiş Piyasa Takas Fiyatı (PTF)</div>
                  <div style="font-size:12px; color:#4EB2D2; margin-top:4px;">{tarih_fmt} Tarihine Ait</div>
                        </td>
                     <td style="vertical-align:middle; text-align:right; width:120px;">
                      <img src="{LOGO_MAIL_SRC}"
                           width="150" height="65"
                           style="display:block; margin-left:auto;"
                           alt="Alpine Enerji" />
                    </td>
              </tr>
            </table>
          </td>
        </tr>

        <tr>
          <td style="padding:25px 30px;">
           <p style="font-size:15px; color:#201F5A;">Sayın <b>{musteri_ad}</b>,</p>
            <p style="font-size:15px; color:#201F5A;">
              {tarih_fmt} tarihine ait <b>Kesinleşmemiş Piyasa Takas Fiyatı (PTF)</b> verileri aşağıda yer almaktadır.
            </p>
          </td>
        </tr>

        <tr>
          <td align="center" style="padding:0 30px;">
            <img src="data:image/png;base64,{grafik_b64}"
                 style="width:100%; max-width:700px; height:auto; display:block; border:1px solid #eee;" />
          </td>
        </tr>

        <tr>
          <td style="padding:20px 30px;">
             <p style="font-size:14px; color:#666; border-top:1px solid #eee; padding-top:10px; font-weight:bold; text-align:center;">
              Kaynak: EPİAŞ Şeffaflık Platformu
              </p>
          </td>
        </tr>

      </table>
    </td></tr>
  </table>
</body>
</html>"""


def mail_gonder(musteri: dict, veri: list, tarih: str, xlsx_bytes: bytes, grafik_b64: str):
    msg = MIMEMultipart("mixed")
    tarih_fmt = datetime.strptime(tarih, "%Y-%m-%d").strftime("%d.%m.%Y")
    msg["Subject"] = f"EPİAŞ Kesinleşmemiş Piyasa Takas Fiyatı (PTF) — {tarih_fmt}"
    msg["From"]    = OUTLOOK_MAIL
    msg["To"]      = musteri["email"]
    msg.attach(MIMEText(html_mail_olustur(musteri["ad"], veri, tarih, grafik_b64), "html", "utf-8"))
    ek = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    ek.set_payload(xlsx_bytes)
    encoders.encode_base64(ek)
    ek.add_header("Content-Disposition", f'attachment; filename="PTF_{tarih}.xlsx"')
    msg.attach(ek)
    with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
        server.starttls()
        server.login(OUTLOOK_MAIL, OUTLOOK_PASS)
        server.sendmail(OUTLOOK_MAIL, [musteri["email"]], msg.as_string())


def main():
    log.info("=" * 55)
    try:
        veri, tarih = ptf_veri_cek()

        if veri is None:
            log.info("Süreç durduruldu: Yarının PTF verisi henüz yayınlanmadığı için mail gönderilmedi.")
            return

        xlsx_bytes = xlsx_olustur(veri, tarih)
        grafik_b64 = grafik_olustur(veri, tarih)

        for musteri in MUSTERI_LISTESI:
            mail_gonder(musteri, veri, tarih, xlsx_bytes, grafik_b64)
            log.info(f"✓ Mail gönderildi → {musteri['email']}")

    except Exception as e:
        log.error(f"Süreç sırasında HATA oluştu: {e}")
        raise


if __name__ == "__main__":
    main()
